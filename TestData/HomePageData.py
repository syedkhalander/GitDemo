import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"Shetty","gender":"Male"},{"firstname":"Syed","lastname":"Khalander","gender":"Male"}]

    @staticmethod
    def getTestData(Testcase):
        book = openpyxl.load_workbook("C:\\Users\\syed khalander\\Documents\\PythonDemo.xlsx")

        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == Testcase:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]